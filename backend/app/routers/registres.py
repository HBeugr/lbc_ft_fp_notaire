"""Registres légaux — 6 registres basés sur audit_logs + export PDF (Art. 23, Ordonnance 2023-875)."""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.audit import AuditLog
from app.models.dossier import Dossier
from app.models.user import User
from app.repositories import audit_repo

router = APIRouter(prefix="/registres", tags=["registres"])

# CDC §7.3 — REG-01 « Consulter les registres » et REG-02 « Exporter les registres » :
# Admin O, Notaire Principal O, Resp. Conformité O, Clercs **N**. Les registres sont
# des pièces de conformité opposables (Art. 23) : leur consultation relève de la
# supervision, pas de la saisie. Aucun registre n'est donc ouvert aux clercs, y
# compris les registres non marqués « confidentiels ».
_ROLES_REGISTRES = ["admin", "notaire_principal", "responsable_conformite"]

REGISTRE_DEFS: dict[str, dict] = {
    "kyc": {
        "label": "Registre KYC",
        "filter_actions": ["dossier.created", "kyc.pp.saved", "kyc.pm.saved"],
        "confidential": False,
        "roles": None,
    },
    "alertes": {
        "label": "Registre des Alertes",
        "filter_actions": ["alerte.created", "alerte.traitee"],
        "confidential": False,
        "roles": None,
    },
    "risque_eleve": {
        "label": "Registre des Dossiers à Risque Élevé (score >= 14)",
        "source": "dossiers",
        "confidential": False,
        "roles": ["admin", "notaire_principal", "responsable_conformite"],
    },
    "dos": {
        "label": "Registre DOS (confidentiel — Art. 63)",
        "filter_actions": ["dos.created", "dos.soumis", "dos.accuse_recu", "dos.addendum"],
        "confidential": True,
        "roles": ["admin", "notaire_principal", "responsable_conformite"],
    },
    "statuts": {
        "label": "Registre des Changements de Statut",
        "filter_actions": ["dossier.statut_change"],
        "confidential": False,
        "roles": None,
    },
    "revisions": {
        "label": "Registre des Révisions KYC",
        "filter_actions": ["revision.created", "revision.validee"],
        "confidential": False,
        "roles": None,
    },
    "journal": {
        "label": "Journal des Actions Utilisateurs",
        "filter_actions": None,
        # Le Notaire Principal était exclu du journal des actions, alors que le CDC
        # lui accorde REG-01 (consulter les registres) ET ADM-06 (consulter les logs).
        # Or c'est l'assujetti au sens de l'Ordonnance : lui refuser la piste d'audit
        # de son propre cabinet le privait du moyen d'exercer sa supervision (Art. 12).
        "confidential": False,
        "roles": ["admin", "notaire_principal", "responsable_conformite"],
    },
}


def _roles_autorises(reg: dict) -> list[str]:
    """Rôles admis sur un registre — repli sur la règle REG-01 (jamais « tout le monde »).

    Un `roles: None` signifiait auparavant « aucune restriction », ce qui ouvrait
    les registres KYC, alertes, statuts et révisions aux clercs en violation de
    REG-01. Le repli est désormais restrictif : c'est la valeur par défaut qui doit
    être sûre, pas l'exception.
    """
    return reg.get("roles") or _ROLES_REGISTRES


def _require_registre_access(reg_id: str, actor: User) -> None:
    reg = REGISTRE_DEFS.get(reg_id)
    if not reg:
        raise HTTPException(status_code=404, detail="Registre inconnu.")
    if not actor.a_role(*_roles_autorises(reg)):
        raise HTTPException(status_code=403, detail="Accès non autorisé à ce registre.")


def _s(text: str) -> str:
    return (text or "").replace("—", "-").replace("–", "-").encode("latin-1", errors="replace").decode("latin-1")


def _generate_pdf(label: str, items: list[AuditLog]) -> bytes:
    from fpdf import FPDF
    from datetime import datetime, timezone

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_fill_color(26, 46, 74)
    pdf.set_text_color(232, 184, 75)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 10, _s(label), ln=True, fill=True, align="C")
    pdf.set_text_color(100, 116, 139)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 5, f"Exporte le {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC", ln=True, align="C")
    pdf.set_text_color(220, 38, 38)
    pdf.cell(0, 5, "CONFIDENTIEL - Usage interne - Art. 23 (conservation 10 ans)", ln=True, align="C")
    pdf.ln(3)

    headers = ["Date/Heure", "Action", "Entite", "ID Entite", "Utilisateur", "IP"]
    col_w = [32, 45, 25, 35, 35, 18]
    pdf.set_fill_color(26, 46, 74)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for w, h in zip(col_w, headers):
        pdf.cell(w, 7, h, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for i, e in enumerate(items):
        pdf.set_fill_color(248, 250, 252) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(30, 41, 59)
        ts = e.created_at.strftime("%d/%m/%Y %H:%M") if e.created_at else "-"
        row = [ts, (e.action or "")[:40], e.entity_type or "-", (e.entity_id or "-")[:30], (e.user_id or "-")[:30], e.ip_address or "-"]
        for w, v in zip(col_w, row):
            pdf.cell(w, 5, _s(str(v))[:38], fill=True)
        pdf.ln()

    return bytes(pdf.output())


def _generate_pdf_dossiers(label: str, rows: list) -> bytes:
    from fpdf import FPDF
    from datetime import datetime, timezone

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_fill_color(26, 46, 74)
    pdf.set_text_color(232, 184, 75)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 10, _s(label), ln=True, fill=True, align="C")
    pdf.set_text_color(100, 116, 139)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 5, f"Exporte le {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC", ln=True, align="C")
    pdf.ln(3)

    headers = ["Reference", "Client", "Operation", "Score", "Classif.", "Trigger", "Statut"]
    col_w = [38, 20, 38, 16, 22, 20, 28]
    pdf.set_fill_color(26, 46, 74)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for w, h in zip(col_w, headers):
        pdf.cell(w, 7, h, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for i, d in enumerate(rows):
        pdf.set_fill_color(248, 250, 252) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(30, 41, 59)
        row = [d.reference or "-", d.type_client or "-", d.type_operation or "-",
               str(d.score_base if d.score_base is not None else "-"),
               d.classification or "-", d.trigger_actif or "-", d.statut or "-"]
        for w, v in zip(col_w, row):
            pdf.cell(w, 5, _s(str(v))[:34], fill=True)
        pdf.ln()

    return bytes(pdf.output())


def _generate_excel(label: str, headers: list[str], data_rows: list[list]) -> bytes:
    """Export Excel (openpyxl) d'un registre — parité avec l'immo."""
    import io
    from datetime import datetime, timezone
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registre"
    ws.append([label])
    ws.append([f"Exporté le {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC — "
               "CONFIDENTIEL, usage interne (Art. 23, conservation 10 ans)"])
    ws.append([])
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="1A2E4A")
    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
    for r in data_rows:
        ws.append([str(v) if v is not None else "" for v in r])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 50)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.get("")
async def list_registres(current_user: User = Depends(get_current_user)) -> dict:
    visible = {}
    for key, reg in REGISTRE_DEFS.items():
        # Même règle que `_require_registre_access` : la liste ne doit pas révéler
        # l'existence de registres que l'appelant n'a pas le droit de consulter.
        if not current_user.a_role(*_roles_autorises(reg)):
            continue
        visible[key] = {"id": key, "label": reg["label"], "confidential": reg.get("confidential", False)}
    return {"registres": list(visible.values())}


@router.get("/{registre_id}")
async def get_registre(
    registre_id: str,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    _require_registre_access(registre_id, current_user)
    reg = REGISTRE_DEFS[registre_id]

    # Registre des dossiers à risque élevé — source = table dossiers (score >= 14 / classification ÉLEVÉ)
    if reg.get("source") == "dossiers":
        base = select(Dossier).where(
            (Dossier.classification == "ELEVE") | (Dossier.score_base >= 14)
        ).order_by(Dossier.updated_at.desc())
        total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
        rows = (await db.execute(base.limit(limit).offset(offset))).scalars().all()
        return {
            "id": registre_id,
            "label": reg["label"],
            "total": total,
            "items": [
                {
                    "id": d.id,
                    "reference": d.reference,
                    "type_client": d.type_client,
                    "type_operation": d.type_operation,
                    "score_base": d.score_base,
                    "classification": d.classification,
                    "trigger_actif": d.trigger_actif,
                    "statut": d.statut,
                    "updated_at": d.updated_at.isoformat() if d.updated_at else None,
                }
                for d in rows
            ],
        }

    q = select(AuditLog).order_by(AuditLog.created_at.desc())
    if reg.get("filter_actions"):
        q = q.where(AuditLog.action.in_(reg["filter_actions"]))

    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar_one()
    result = await db.execute(q.limit(limit).offset(offset))
    items = result.scalars().all()

    return {
        "id": registre_id,
        "label": reg["label"],
        "total": total,
        "items": [
            {
                "id": e.id,
                "timestamp": e.created_at.isoformat() if e.created_at else None,
                "action": e.action,
                "entity_type": e.entity_type,
                "entity_id": e.entity_id,
                "user_id": e.user_id,
                "user_role": e.user_role,
                "ip": e.ip_address,
                "detail": e.detail,
            }
            for e in items
        ],
    }


@router.get("/{registre_id}/export")
async def export_registre(
    registre_id: str,
    format: str = Query("pdf", regex="^(pdf|excel)$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> Response:
    _require_registre_access(registre_id, current_user)
    reg = REGISTRE_DEFS[registre_id]
    from datetime import datetime, timezone

    if reg.get("source") == "dossiers":
        rows = list((await db.execute(
            select(Dossier).where(
                (Dossier.classification == "ELEVE") | (Dossier.score_base >= 14)
            ).order_by(Dossier.updated_at.desc()).limit(1000)
        )).scalars().all())
        await audit_repo.log(
            db, action=f"registre.{registre_id}.exported", user_id=current_user.id,
            user_role=current_user.role, entity_type="registre", entity_id=registre_id,
            ip="internal", detail={"format": format, "count": len(rows)},
        )
        base = f"registre-{registre_id}-{datetime.now(timezone.utc).strftime('%Y%m%d')}"
        if format == "excel":
            headers = ["Reference", "Client", "Operation", "Score", "Classif.", "Trigger", "Statut"]
            drows = [[d.reference, d.type_client, d.type_operation,
                      d.score_base if d.score_base is not None else "",
                      d.classification, d.trigger_actif, d.statut] for d in rows]
            return Response(content=_generate_excel(reg["label"], headers, drows), media_type=_XLSX_MIME,
                            headers={"Content-Disposition": f'attachment; filename="{base}.xlsx"'})
        return Response(content=_generate_pdf_dossiers(reg["label"], rows), media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="{base}.pdf"'})

    q = select(AuditLog).order_by(AuditLog.created_at.desc()).limit(1000)
    if reg.get("filter_actions"):
        q = q.where(AuditLog.action.in_(reg["filter_actions"]))
    result = await db.execute(q)
    items = list(result.scalars().all())

    await audit_repo.log(
        db,
        action=f"registre.{registre_id}.exported",
        user_id=current_user.id,
        user_role=current_user.role,
        entity_type="registre",
        entity_id=registre_id,
        ip="internal",
        detail={"format": format, "count": len(items)},
    )

    base = f"registre-{registre_id}-{datetime.now(timezone.utc).strftime('%Y%m%d')}"
    if format == "excel":
        headers = ["Date/Heure", "Action", "Entite", "ID Entite", "Utilisateur", "IP"]
        drows = [[e.created_at.strftime("%d/%m/%Y %H:%M") if e.created_at else "",
                  e.action, e.entity_type, e.entity_id, e.user_id, e.ip_address] for e in items]
        return Response(content=_generate_excel(reg["label"], headers, drows), media_type=_XLSX_MIME,
                        headers={"Content-Disposition": f'attachment; filename="{base}.xlsx"'})
    return Response(content=_generate_pdf(reg["label"], items), media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{base}.pdf"'})
