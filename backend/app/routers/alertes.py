import asyncio
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.core.database import get_db, tenant_session
from app.core.deps import get_current_user, require_rc
from app.core import security
from app.core.tenant_context import get_current_tenant, tenant_scope
from pydantic import BaseModel
from app.models.alerte import Alerte
from app.models.user import User
from app.repositories import alertes_repo, audit_repo, dossier_repo, user_repo
from app.schemas.alertes import AlerteCreate, AlerteOut, AlerteTraiter, AlerteListResponse, SignalementInterneRequest


class AlerteTraiterPost(BaseModel):
    """Contrat frontend : traitement d'une alerte avec justification + action optionnelle sur le dossier lié."""
    justification: str
    action_dossier: str | None = None


_ACTIONS_DOSSIER = {"AUCUNE", "DEBLOQUER", "BLOQUER", "DEMANDER_DOCUMENTS", "DECLENCHER_DOS"}

router = APIRouter(prefix="/alertes", tags=["alertes"])


@router.get("", response_model=AlerteListResponse)
async def list_alertes(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    statut: str | None = Query(None),
    niveau: str | None = Query(None),
    type_alerte: str | None = Query(None),
    dossier_id: str | None = Query(None),
    categorie: str | None = Query(None, description="conformite | notification | historique"),
    dossier_statut: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
) -> AlerteListResponse:
    # Le frontend envoie le statut en MAJUSCULES (OUVERTE/…) ; la base stocke en minuscules.
    statut_db = statut.lower() if statut else None
    # Cloisonnement Art.63 : un non-superviseur ne voit que les alertes de SES dossiers assignés.
    scope_assigned = None if current_user.is_supervisor else current_user.id
    items, total = await alertes_repo.list_alertes(
        db,
        statut=statut_db,
        niveau=niveau,
        type_alerte=type_alerte,
        dossier_id=dossier_id,
        categorie=categorie,
        dossier_statut=dossier_statut,
        assigned_to=scope_assigned,
        limit=page_size,
        offset=(page - 1) * page_size,
    )
    # Enrichissement référence + statut du dossier lié (fetch groupé, sans lazy-load)
    from app.models.dossier import Dossier
    dossier_ids = {a.dossier_id for a in items if a.dossier_id}
    dossiers_map: dict[str, tuple[str, str]] = {}
    if dossier_ids:
        rows = (await db.execute(
            select(Dossier.id, Dossier.reference, Dossier.statut).where(Dossier.id.in_(dossier_ids))
        )).all()
        dossiers_map = {r.id: (r.reference, r.statut) for r in rows}
    out = []
    for a in items:
        ref, dstatut = dossiers_map.get(a.dossier_id, (None, None))
        out.append(AlerteOut.from_orm_safe(a, dossier_reference=ref, dossier_statut=dstatut))
    return AlerteListResponse(
        items=out,
        total=total,
        page=page,
        page_size=page_size,
    )


@router.post("", response_model=AlerteOut, status_code=status.HTTP_201_CREATED)
async def create_alerte(
    body: AlerteCreate,
    request: Request,
    current_user: User = Depends(require_rc),
    db: AsyncSession = Depends(get_db),
) -> AlerteOut:
    alerte = await alertes_repo.create(
        db,
        dossier_id=body.dossier_id,
        type_alerte=body.type_alerte,
        niveau=body.niveau,
        description=body.description,
    )
    ip = request.client.host if request.client else "unknown"
    await audit_repo.log(
        db,
        action="alerte.created",
        user_id=current_user.id,
        user_role=current_user.role,
        entity_type="alerte",
        entity_id=alerte.id,
        ip=ip,
        detail={"type_alerte": body.type_alerte, "niveau": body.niveau, "dossier_id": body.dossier_id},
    )
    return AlerteOut.from_orm_safe(alerte)


# ── Signalement interne (clercs) — doit précéder /{alerte_id} ─────────────────
_SIGNALEUR_ROLES = ("clercs", "admin")


@router.post("/signaler", response_model=AlerteOut, status_code=status.HTTP_201_CREATED)
async def signaler_alerte_interne(
    body: SignalementInterneRequest,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> AlerteOut:
    """Un clerc signale une suspicion au Responsable Conformité (CDC — Art. 29)."""
    if current_user.role not in _SIGNALEUR_ROLES:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Accès réservé aux clercs et à l'administrateur.",
        )
    dossier_id: str | None = None
    if body.dossier_reference:
        dossier = await dossier_repo.get_by_reference(db, body.dossier_reference)
        if dossier is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Dossier introuvable : {body.dossier_reference}",
            )
        dossier_id = dossier.id
    alerte = await alertes_repo.create(
        db,
        dossier_id=dossier_id,
        type_alerte="SIGNALEMENT_INTERNE",
        niveau="MOYEN",
        description=f"[Signalement {current_user.first_name} {current_user.last_name}] {body.description}",
        signaleur_id=current_user.id,
    )
    ip = request.client.host if request.client else "unknown"
    await audit_repo.log(
        db,
        action="alerte.signalement_interne",
        user_id=current_user.id,
        user_role=current_user.role,
        entity_type="alerte",
        entity_id=alerte.id,
        ip=ip,
        detail={"dossier_reference": body.dossier_reference},
    )
    return AlerteOut.from_orm_safe(alerte)


@router.get("/mes-signalements", response_model=list[AlerteOut])
async def mes_signalements(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[AlerteOut]:
    """Retourne les signalements internes émis par l'utilisateur courant."""
    if current_user.role not in _SIGNALEUR_ROLES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Accès réservé aux clercs.")
    items = await alertes_repo.list_by_signaleur(db, signaleur_id=current_user.id)
    return [AlerteOut.from_orm_safe(a) for a in items]


# ── Temps réel : compteur d'alertes (badge SSE) — doit précéder /{alerte_id} ──

@router.get("/mon-compteur")
async def mon_compteur(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Compteur d'alertes non traitées pour l'utilisateur (badge — fallback du SSE)."""
    count = await alertes_repo.count_open_for_user(db, current_user.id, current_user.is_supervisor)
    return {"count": count}


@router.get("/stream")
async def alertes_stream(token: str = Query(...)):
    """Flux SSE temps réel du compteur d'alertes (badge).
    Auth par query param car EventSource ne supporte pas l'en-tête Authorization."""
    from app.core.redis_client import is_token_revoked, is_user_globally_revoked
    try:
        payload = security.decode_token(token)
        user_id = payload.get("sub")
    except Exception:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token invalide.")
    # Durcissement : refuser tout token non-access (ex. refresh) ou révoqué.
    if not user_id or payload.get("type") != "access":
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token invalide.")
    jti = payload.get("jti")
    if (jti and await is_token_revoked(jti)) or await is_user_globally_revoked(user_id):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Session révoquée.")

    # Le cabinet doit être capturé ICI, dans le corps de l'endpoint : le générateur
    # SSE est consommé APRÈS la sortie du middleware, hors du `ContextVar` de
    # cloisonnement. Sans cette capture, chaque session du flux tape le schéma par
    # défaut (« relation "users" does not exist ») et le badge reste muet.
    tenant = get_current_tenant()

    async def _in_tenant(run):
        """Exécute une lecture dans le cabinet capturé, sur une session cloisonnée.

        Le `tenant_scope` est refermé AVANT tout `yield` du générateur : un
        `ContextVar` posé à cheval sur une suspension fuiterait dans la tâche
        qui consomme le flux.
        """
        with tenant_scope(tenant):
            async with tenant_session(tenant) as session:
                return await run(session)

    async def event_gen():
        # Charge le rôle une fois (is_supervisor) pour le périmètre du compteur
        try:
            user = await _in_tenant(lambda s: user_repo.get_by_id(s, user_id))
        except Exception:
            user = None
        if user is None:
            yield ": error\n\n"
            return
        is_sup = user.is_supervisor
        last = None
        while True:
            try:
                count = await _in_tenant(
                    lambda s: alertes_repo.count_open_for_user(s, user_id, is_sup)
                )
                if count != last:
                    last = count
                    yield f"event: count\ndata: {count}\n\n"
                else:
                    yield ": keepalive\n\n"
            except Exception:
                yield ": error\n\n"
            await asyncio.sleep(3)

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/export")
async def export_alertes(
    request: Request,
    format: str = Query("excel", pattern="^(excel|pdf)$"),
    statut: str | None = Query(None),
    niveau: str | None = Query(None),
    type_alerte: str | None = Query(None),
    categorie: str | None = Query(None),
    dossier_statut: str | None = Query(None),
    current_user: User = Depends(require_rc),
    db: AsyncSession = Depends(get_db),
):
    """Export du registre des alertes (Excel ou PDF) — réservé conformité."""
    from app.models.dossier import Dossier

    statut_db = statut.lower() if statut else None
    items, _ = await alertes_repo.list_alertes(
        db, statut=statut_db, niveau=niveau, type_alerte=type_alerte,
        categorie=categorie, dossier_statut=dossier_statut, limit=10_000, offset=0,
    )
    # Traçabilité LBC/FT : l'export du registre des alertes est journalisé.
    await audit_repo.log(
        db, action="alerte.export", user_id=current_user.id, user_role=current_user.role,
        entity_type="alerte", entity_id="export",
        ip=request.client.host if request.client else "unknown",
        detail={"format": format, "categorie": categorie, "lignes": len(items)},
    )
    # Références des dossiers liés (fetch groupé)
    dossier_ids = {a.dossier_id for a in items if a.dossier_id}
    refs: dict[str, str] = {}
    if dossier_ids:
        rows = (await db.execute(
            select(Dossier.id, Dossier.reference).where(Dossier.id.in_(dossier_ids))
        )).all()
        refs = {r.id: r.reference for r in rows}

    headers = ["Dossier", "Type", "Niveau", "Statut", "Description", "Créée le", "Traitée le", "Traitée par"]

    def _row(a) -> list[str]:
        return [
            refs.get(a.dossier_id, a.dossier_id or "—"),
            a.type_alerte, a.niveau, a.statut, (a.description or "")[:200],
            a.created_at.strftime("%d/%m/%Y %H:%M") if a.created_at else "",
            a.traite_at.strftime("%d/%m/%Y %H:%M") if a.traite_at else "",
            a.traite_par or "",
        ]

    if format == "excel":
        from io import BytesIO
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Alertes"
        ws.append(headers)
        for a in items:
            ws.append(_row(a))
        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=alertes.xlsx"},
        )

    # PDF
    from fpdf import FPDF
    pdf = FPDF(orientation="L")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 10, "Registre des alertes", ln=1)
    pdf.set_font("Helvetica", "", 7)
    widths = [32, 38, 18, 20, 80, 30, 30, 30]
    for h, w in zip(headers, widths):
        pdf.cell(w, 7, h, border=1)
    pdf.ln()
    for a in items:
        for val, w in zip(_row(a), widths):
            pdf.cell(w, 6, str(val)[:60].encode("latin-1", "replace").decode("latin-1"), border=1)
        pdf.ln()
    out = pdf.output(dest="S")
    pdf_bytes = out.encode("latin-1") if isinstance(out, str) else bytes(out)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=alertes.pdf"},
    )


async def _assert_alerte_read_access(db: AsyncSession, alerte, user: User) -> None:
    """Cloisonnement Art.63 : un non-superviseur ne lit qu'une alerte de SES dossiers assignés."""
    if user.is_supervisor:
        return
    if not alerte.dossier_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Accès refusé.")
    dossier = await dossier_repo.get_by_id(db, alerte.dossier_id)
    if not dossier or dossier.assigned_to != user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Accès refusé.")


@router.get("/{alerte_id}/timeline")
async def alerte_timeline(
    alerte_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Parcours chronologique d'une alerte (créée → traitée + note)."""
    result = await db.execute(select(Alerte).where(Alerte.id == alerte_id))
    a = result.scalar_one_or_none()
    if not a:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Alerte introuvable.")
    await _assert_alerte_read_access(db, a, current_user)
    events = [{"label": "Alerte créée", "at": a.created_at.isoformat() if a.created_at else None, "par": None}]
    if a.prise_en_charge_at:
        events.append({
            "label": "Prise en charge",
            "at": a.prise_en_charge_at.isoformat(),
            "par": a.prise_en_charge_par,
        })
    if a.traite_at:
        events.append({
            "label": "Traitée",
            "at": a.traite_at.isoformat(),
            "par": a.traite_par,
            "note": a.resolution_note,
        })
    return {"alerte_id": alerte_id, "statut": a.statut, "type_alerte": a.type_alerte, "events": events}


@router.post("/{alerte_id}/prendre", response_model=AlerteOut)
async def prendre_alerte(
    alerte_id: str,
    request: Request,
    current_user: User = Depends(require_rc),
    db: AsyncSession = Depends(get_db),
) -> AlerteOut:
    """Prise en charge d'une alerte (ouverte → en_cours), traçabilité qui/quand."""
    result = await db.execute(select(Alerte).where(Alerte.id == alerte_id))
    alerte = result.scalar_one_or_none()
    if not alerte:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Alerte introuvable.")
    if alerte.statut != "ouverte":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Seule une alerte ouverte peut être prise en charge.")
    alerte.statut = "en_cours"
    alerte.prise_en_charge_par = current_user.id
    alerte.prise_en_charge_at = datetime.now(timezone.utc)
    db.add(alerte)
    await db.commit()
    await db.refresh(alerte)
    ip = request.client.host if request.client else "unknown"
    await audit_repo.log(
        db, action="alerte.prise_en_charge", user_id=current_user.id, user_role=current_user.role,
        entity_type="alerte", entity_id=alerte_id, ip=ip,
    )
    return AlerteOut.from_orm_safe(alerte)


@router.get("/{alerte_id}", response_model=AlerteOut)
async def get_alerte(
    alerte_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> AlerteOut:
    result = await db.execute(select(Alerte).where(Alerte.id == alerte_id))
    alerte = result.scalar_one_or_none()
    if not alerte:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Alerte introuvable.")
    await _assert_alerte_read_access(db, alerte, current_user)
    return AlerteOut.from_orm_safe(alerte)


@router.patch("/{alerte_id}/traiter", response_model=AlerteOut)
async def traiter_alerte(
    alerte_id: str,
    body: AlerteTraiter,
    request: Request,
    current_user: User = Depends(require_rc),
    db: AsyncSession = Depends(get_db),
) -> AlerteOut:
    result = await db.execute(select(Alerte).where(Alerte.id == alerte_id))
    alerte = result.scalar_one_or_none()
    if not alerte:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Alerte introuvable.")
    alerte = await alertes_repo.update_statut(db, alerte, body.statut, current_user.id, body.resolution_note)
    # Consolidation : refermer les doublons actifs du même type sur le dossier (ex. plusieurs
    # T3_SANCTIONS/RCCM_EXPIRE rejoués à chaque sauvegarde) → une trace traitée unique par type.
    doublons_consolides = 0
    if body.statut == "traitee" and alerte.dossier_id:
        doublons_consolides = await alertes_repo.resoudre_actives_par_type(
            db, dossier_id=alerte.dossier_id, type_alerte=alerte.type_alerte,
            traite_par=current_user.id, note="Doublon consolidé au traitement.",
        )
        if doublons_consolides:
            await db.commit()
    ip = request.client.host if request.client else "unknown"
    await audit_repo.log(
        db,
        action="alerte.traitee",
        user_id=current_user.id,
        user_role=current_user.role,
        entity_type="alerte",
        entity_id=alerte_id,
        ip=ip,
        detail={"statut": body.statut, "note": body.resolution_note, "doublons_consolides": doublons_consolides},
    )
    return AlerteOut.from_orm_safe(alerte)


@router.post("/{alerte_id}/traiter", response_model=AlerteOut)
async def traiter_alerte_post(
    alerte_id: str,
    body: AlerteTraiterPost,
    request: Request,
    current_user: User = Depends(require_rc),
    db: AsyncSession = Depends(get_db),
) -> AlerteOut:
    """Traite une alerte (justification obligatoire) + couplage optionnel action→KYC
    (BLOQUER/DEBLOQUER le dossier lié). Contrat frontend AlertesView."""
    result = await db.execute(select(Alerte).where(Alerte.id == alerte_id))
    alerte = result.scalar_one_or_none()
    if not alerte:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Alerte introuvable.")
    if alerte.statut == "traitee":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Cette alerte est déjà traitée.")
    action = body.action_dossier
    if action and action not in _ACTIONS_DOSSIER:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Action dossier invalide : {action}.")

    ip = request.client.host if request.client else "unknown"
    # Couplage alerte → KYC : applique l'action choisie sur le dossier lié.
    # DECLENCHER_DOS bloque le dossier (l'ouverture d'une DOS bloque le dossier) — la création
    # de la DOS elle-même se fait ensuite via le formulaire DOS (navigation côté frontend).
    if action in ("BLOQUER", "DEBLOQUER", "DECLENCHER_DOS") and alerte.dossier_id:
        dossier = await dossier_repo.get_by_id(db, alerte.dossier_id)
        if dossier:
            if action == "DEBLOQUER":
                dossier.statut, dossier.is_bloque = "en_analyse", False
            else:  # BLOQUER ou DECLENCHER_DOS
                dossier.statut, dossier.is_bloque = "bloque", True
            db.add(dossier)
            await db.commit()
            await audit_repo.log(
                db, action=f"dossier.action_depuis_alerte.{action.lower()}",
                user_id=current_user.id, user_role=current_user.role,
                entity_type="dossier", entity_id=dossier.id, ip=ip,
                detail={"nouveau_statut": dossier.statut},
            )

    alerte = await alertes_repo.update_statut(db, alerte, "traitee", current_user.id, body.justification)
    doublons = 0
    if alerte.dossier_id:
        doublons = await alertes_repo.resoudre_actives_par_type(
            db, dossier_id=alerte.dossier_id, type_alerte=alerte.type_alerte,
            traite_par=current_user.id, note="Doublon consolidé au traitement.",
        )
        if doublons:
            await db.commit()
    await audit_repo.log(
        db, action="alerte.traitee", user_id=current_user.id, user_role=current_user.role,
        entity_type="alerte", entity_id=alerte_id, ip=ip,
        detail={"justification": body.justification, "action_dossier": action, "doublons_consolides": doublons},
    )
    return AlerteOut.from_orm_safe(alerte)


async def _dossier_from_alerte(db: AsyncSession, alerte_id: str):
    result = await db.execute(select(Alerte).where(Alerte.id == alerte_id))
    alerte = result.scalar_one_or_none()
    if not alerte:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Alerte introuvable.")
    if not alerte.dossier_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cette alerte n'est liée à aucun dossier.")
    dossier = await dossier_repo.get_by_id(db, alerte.dossier_id)
    if not dossier:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dossier introuvable.")
    return dossier


@router.post("/{alerte_id}/bloquer-dossier")
async def bloquer_dossier_depuis_alerte(
    alerte_id: str,
    request: Request,
    current_user: User = Depends(require_rc),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Bloque le dossier lié à l'alerte (ALE03)."""
    dossier = await _dossier_from_alerte(db, alerte_id)
    if dossier.statut == "bloque":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Le dossier est déjà bloqué.")
    dossier.statut, dossier.is_bloque = "bloque", True
    db.add(dossier)
    await db.commit()
    ip = request.client.host if request.client else "unknown"
    await audit_repo.log(
        db, action="dossier.bloque_depuis_alerte", user_id=current_user.id, user_role=current_user.role,
        entity_type="dossier", entity_id=dossier.id, ip=ip, detail={"new_statut": "bloque"},
    )
    return {"status": "bloque", "dossier_id": dossier.id}


@router.post("/{alerte_id}/debloquer-dossier")
async def debloquer_dossier_depuis_alerte(
    alerte_id: str,
    request: Request,
    current_user: User = Depends(require_rc),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Débloque le dossier lié à l'alerte (ALE03)."""
    dossier = await _dossier_from_alerte(db, alerte_id)
    if dossier.statut != "bloque":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Le dossier n'est pas bloqué.")
    dossier.statut, dossier.is_bloque = "en_analyse", False
    db.add(dossier)
    await db.commit()
    ip = request.client.host if request.client else "unknown"
    await audit_repo.log(
        db, action="dossier.debloque_depuis_alerte", user_id=current_user.id, user_role=current_user.role,
        entity_type="dossier", entity_id=dossier.id, ip=ip, detail={"new_statut": "en_analyse"},
    )
    return {"status": "en_analyse", "dossier_id": dossier.id}
